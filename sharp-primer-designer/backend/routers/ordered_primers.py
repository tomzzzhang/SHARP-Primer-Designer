"""CRUD + import endpoints for the ordered-primers exclusion library.

The library is a list of primer 5'->3' sequences that have already been ordered.
The design pipeline drops any candidate pair where either primer (case-insensitive)
matches an entry, preventing accidental re-ordering of primers already in inventory.
"""

from __future__ import annotations

import io
import json
import re
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, File, HTTPException, UploadFile

from core.models import (
    BulkOrderedPrimersRequest,
    BulkOrderedPrimersResponse,
    OrderedPrimer,
    OrderedPrimersResponse,
)

router = APIRouter(prefix="/api/ordered_primers", tags=["ordered_primers"])

_ORDERED_PATH = Path(__file__).parent.parent / "data" / "ordered_primers.json"

# Sequence validation: ACGTN only after cleaning. We allow degenerate code "N"
# because IUPAC ambiguity codes can show up in real ordered sequences, but for
# matching against primer3 candidates we only need exact ACGT comparison.
_VALID_SEQ_RE = re.compile(r"^[ACGTN]+$")


# ─── Persistence ──────────────────────────────────────────────────────────────


def _load() -> list[OrderedPrimer]:
    if not _ORDERED_PATH.exists():
        return []
    with open(_ORDERED_PATH) as f:
        data = json.load(f)
    return [OrderedPrimer(**p) for p in data.get("primers", [])]


def _save(primers: list[OrderedPrimer]) -> None:
    with open(_ORDERED_PATH, "w") as f:
        json.dump({"primers": [p.model_dump() for p in primers]}, f, indent=2)


# ─── Sequence parsing helpers ─────────────────────────────────────────────────


def _normalize_sequence(seq: str) -> str:
    """Uppercase + strip whitespace + drop non-letter chars. Returns '' if empty."""
    return re.sub(r"[^A-Za-z]", "", seq).upper()


def _parse_pasted_sequences(raw_lines: list[str]) -> list[str]:
    """Parse pasted text. Accepts either:
       - One sequence per line (FASTA headers starting with '>' separate records)
       - Multi-FASTA blocks
       - Mixed (any line starting with '>' is a header; subsequent non-header
         lines accumulate into one sequence until the next header)
    Returns a list of uppercased ACGT-only sequences (empty entries dropped).
    """
    sequences: list[str] = []
    current_buf: list[str] = []

    def _flush():
        if current_buf:
            joined = "".join(current_buf)
            cleaned = _normalize_sequence(joined)
            if cleaned:
                sequences.append(cleaned)
            current_buf.clear()

    has_fasta_header = any(line.lstrip().startswith(">") for line in raw_lines)

    if has_fasta_header:
        # Multi-FASTA mode: '>' lines separate records
        for line in raw_lines:
            stripped = line.strip()
            if stripped.startswith(">"):
                _flush()
            elif stripped:
                current_buf.append(stripped)
        _flush()
    else:
        # One-per-line mode
        for line in raw_lines:
            cleaned = _normalize_sequence(line)
            if cleaned:
                sequences.append(cleaned)

    return sequences


def _extract_sequences_from_notion_record(record: dict) -> list[str]:
    """Pull primer sequences out of an exported Notion record JSON.

    The export format (see `routers/export.py:_build_notion_record`) stores oligos
    under `record["oligos"]`, each entry has `entry["Sequence (5->3)"]`.
    """
    oligos = record.get("oligos") or []
    seqs: list[str] = []
    for oligo in oligos:
        if not isinstance(oligo, dict):
            continue
        entry = oligo.get("entry") or {}
        seq = entry.get("Sequence (5->3)") or entry.get("Sequence (5'->3')") or entry.get("Sequence")
        if isinstance(seq, str):
            cleaned = _normalize_sequence(seq)
            if cleaned:
                seqs.append(cleaned)
    return seqs


def _extract_sequences_from_xlsx(content: bytes) -> list[str]:
    """Read primer sequences from an IDT-style .xlsx file.

    Mirrors the column shape produced by `_build_idt_xlsx` (header row:
    Name | Sequence | Scale | Purification). Tolerant of extra columns or
    different casings — finds any column whose header (lowercased, stripped)
    starts with 'sequence'.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(500, f"openpyxl not available: {exc}") from exc

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(422, f"Could not read .xlsx: {exc}") from exc

    seqs: list[str] = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue

        seq_col_idx = None
        for idx, val in enumerate(header):
            if val is None:
                continue
            label = str(val).strip().lower()
            if label.startswith("sequence"):
                seq_col_idx = idx
                break
        if seq_col_idx is None:
            continue

        for row in rows:
            if seq_col_idx >= len(row):
                continue
            cell = row[seq_col_idx]
            if cell is None:
                continue
            cleaned = _normalize_sequence(str(cell))
            if cleaned:
                seqs.append(cleaned)

    return seqs


def _extract_sequences_from_zip(content: bytes) -> list[str]:
    """Walk a zip from this app's export pipeline. Reads the .json inside
    (preferred — has all oligos). Falls back to .xlsx if no .json found."""
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile as exc:
        raise HTTPException(422, f"Not a valid zip file: {exc}") from exc

    json_seqs: list[str] = []
    xlsx_seqs: list[str] = []
    for name in zf.namelist():
        lower = name.lower()
        if lower.endswith("/") or lower.startswith("__macosx"):
            continue
        if lower.endswith(".json"):
            try:
                record = json.loads(zf.read(name).decode("utf-8"))
                json_seqs.extend(_extract_sequences_from_notion_record(record))
            except (json.JSONDecodeError, UnicodeDecodeError):
                continue
        elif lower.endswith(".xlsx"):
            xlsx_seqs.extend(_extract_sequences_from_xlsx(zf.read(name)))

    return json_seqs if json_seqs else xlsx_seqs


# ─── Add helper (shared between bulk + import) ───────────────────────────────


def _add_unique_sequences(
    existing: list[OrderedPrimer],
    new_seqs: list[str],
    source: str,
) -> tuple[list[OrderedPrimer], int, int, list[OrderedPrimer]]:
    """Add new sequences to the library, deduping against existing entries
    and against each other. Returns (updated_list, added_count, skipped_count,
    just_added_entries)."""
    existing_set = {p.sequence.upper() for p in existing}
    added: list[OrderedPrimer] = []
    skipped = 0
    now_iso = datetime.now(timezone.utc).isoformat()

    seen_in_batch: set[str] = set()
    for raw in new_seqs:
        if not _VALID_SEQ_RE.fullmatch(raw):
            skipped += 1
            continue
        if raw in existing_set or raw in seen_in_batch:
            skipped += 1
            continue
        primer = OrderedPrimer(
            id=uuid.uuid4().hex[:12],
            sequence=raw,
            added_date=now_iso,
            source=source,
        )
        added.append(primer)
        seen_in_batch.add(raw)

    updated = existing + added
    return updated, len(added), skipped, added


# ─── Endpoints ────────────────────────────────────────────────────────────────


@router.get("", response_model=OrderedPrimersResponse)
def list_ordered_primers():
    return OrderedPrimersResponse(primers=_load())


@router.post("/bulk", response_model=BulkOrderedPrimersResponse, status_code=201)
def bulk_add_ordered_primers(req: BulkOrderedPrimersRequest):
    """Add many sequences at once. Accepts plain sequences or FASTA blocks
    (one per item or as multi-line text passed in a single item)."""
    if not req.sequences:
        raise HTTPException(400, "No sequences provided")

    raw_lines: list[str] = []
    for entry in req.sequences:
        # Support callers who pass a single string with embedded newlines,
        # OR a list where each element is one sequence.
        raw_lines.extend(entry.splitlines() if "\n" in entry else [entry])

    parsed = _parse_pasted_sequences(raw_lines)
    if not parsed:
        raise HTTPException(400, "No valid sequences found in input")

    existing = _load()
    updated, added_count, skipped_count, added_entries = _add_unique_sequences(
        existing, parsed, source=req.source
    )
    _save(updated)
    return BulkOrderedPrimersResponse(
        added=added_count,
        skipped=skipped_count,
        primers=added_entries,
    )


@router.post("/import", response_model=BulkOrderedPrimersResponse, status_code=201)
async def import_ordered_primers(file: UploadFile = File(...)):
    """Import sequences from a previously-exported zip/json or an IDT-style xlsx."""
    name = (file.filename or "").lower()
    content = await file.read()

    if name.endswith(".zip"):
        seqs = _extract_sequences_from_zip(content)
        source = "imported_json"
    elif name.endswith(".json"):
        try:
            record = json.loads(content.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            raise HTTPException(422, f"Invalid JSON: {exc}") from exc
        seqs = _extract_sequences_from_notion_record(record)
        source = "imported_json"
    elif name.endswith(".xlsx"):
        seqs = _extract_sequences_from_xlsx(content)
        source = "imported_xlsx"
    else:
        raise HTTPException(
            400, f"Unsupported file type: {file.filename}. Use .json, .zip, or .xlsx."
        )

    if not seqs:
        raise HTTPException(422, "No primer sequences found in file")

    existing = _load()
    updated, added_count, skipped_count, added_entries = _add_unique_sequences(
        existing, seqs, source=source
    )
    _save(updated)
    return BulkOrderedPrimersResponse(
        added=added_count,
        skipped=skipped_count,
        primers=added_entries,
    )


@router.delete("/{primer_id}", status_code=204)
def delete_ordered_primer(primer_id: str):
    primers = _load()
    remaining = [p for p in primers if p.id != primer_id]
    if len(remaining) == len(primers):
        raise HTTPException(404, f"Ordered primer '{primer_id}' not found")
    _save(remaining)


@router.delete("", status_code=204)
def clear_ordered_primers(confirm: bool = False):
    if not confirm:
        raise HTTPException(400, "Pass ?confirm=true to clear all ordered primers")
    _save([])
